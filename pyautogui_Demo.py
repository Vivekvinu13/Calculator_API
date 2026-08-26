'''
import pyautogui
import time
import subprocess

# Emergency stop
pyautogui.FAILSAFE = True

# Give yourself time before starting
time.sleep(2)


# --------------------------------------------------
# 1. Open Chrome with Cricbuzz
# --------------------------------------------------

subprocess.run([
    "open",
    "-a",
    "Google Chrome",
    "https://www.cricbuzz.com/"
])

time.sleep(8)


# --------------------------------------------------
# 2. Make Chrome active
# --------------------------------------------------

subprocess.run([
    "osascript",
    "-e",
    'tell application "Google Chrome" to activate'
])

time.sleep(2)


# --------------------------------------------------
# 3. Click inside the webpage
# --------------------------------------------------

pyautogui.click(900, 500)
time.sleep(1)


# --------------------------------------------------
# 4. Select and copy webpage content
# --------------------------------------------------

pyautogui.hotkey("command", "a")
time.sleep(1)

pyautogui.hotkey("command", "c")
time.sleep(2)


# --------------------------------------------------
# 5. Open TextEdit
# --------------------------------------------------

subprocess.run([
    "open",
    "-a",
    "TextEdit"
])

time.sleep(3)


# --------------------------------------------------
# 6. Make TextEdit active
# --------------------------------------------------

subprocess.run([
    "osascript",
    "-e",
    'tell application "TextEdit" to activate'
])

time.sleep(2)


# --------------------------------------------------
# 7. Create a new document
# --------------------------------------------------

pyautogui.hotkey("command", "n")
time.sleep(2)


# --------------------------------------------------
# 8. Paste the webpage content
# --------------------------------------------------

pyautogui.hotkey("command", "v")
time.sleep(3)


# --------------------------------------------------
# 9. Save the document
# --------------------------------------------------

pyautogui.hotkey("command", "s")
time.sleep(3)


# --------------------------------------------------
# 10. Enter filename
# --------------------------------------------------

pyautogui.write("pygui_demo")
time.sleep(1)


# --------------------------------------------------
# 11. Choose Desktop
# --------------------------------------------------

pyautogui.hotkey("command", "shift", "d")
time.sleep(2)


# --------------------------------------------------
# 12. Save
# --------------------------------------------------

pyautogui.press("enter")
time.sleep(4)


print("Demo completed and saved as pygui_demo on Desktop!")
'''
'''
import pyautogui
import time
import subprocess

# Emergency stop
pyautogui.FAILSAFE = True

time.sleep(2)

print("Step 1: Open Chrome...")

pyautogui.hotkey("command", "space", interval=0.1)
time.sleep(1)

pyautogui.write("chrome", interval=0.15)
time.sleep(1)

pyautogui.press("enter")
time.sleep(4)


print("Step 2: Open a new tab...")

pyautogui.hotkey("command", "t", interval=0.1)
time.sleep(2)


print("Step 3: Go to the website...")

pyautogui.hotkey("command", "l", interval=0.1)
time.sleep(1)

pyautogui.typewrite(
    "https://www.accuweather.com/en/in/chennai/206671/current-weather/206671",
    interval=0.03
)
time.sleep(1)

pyautogui.press("enter")
time.sleep(7)


print("Step 4: Select the webpage...")

pyautogui.click(900, 500)

time.sleep(1)

pyautogui.hotkey("command", "a", interval=0.1)
time.sleep(1)

print("Step 5: Copy the webpage...")

pyautogui.hotkey("command", "c", interval=0.1)
time.sleep(1)

print("Step 6: Open TextEdit...")

pyautogui.hotkey("command", "space", interval=0.1)
time.sleep(1)

pyautogui.write("TextEdit", interval=0.15)
time.sleep(1)

pyautogui.press("enter")
time.sleep(4)

# 6. Open TextEdit
# --------------------------------------------------

subprocess.run([
    "open",
    "-a",
    "TextEdit"
])

time.sleep(3)
# 7. Make TextEdit active

subprocess.run([
    "osascript",
    "-e",
    'tell application "TextEdit" to activate'
])

time.sleep(2)

print("Step 7: Create a new document...")

pyautogui.hotkey("command", "n", interval=0.1)
time.sleep(3)

print("Step 8: Paste the copied content...")

pyautogui.hotkey("command", "v", interval=0.1)
time.sleep(3)

print("Step 9: Save the document...")

pyautogui.hotkey("command", "s", interval=0.1)
time.sleep(3)

pyautogui.write("pygui_demo", interval=0.15)
time.sleep(1)

pyautogui.hotkey("command", "shift", "d", interval=0.1)
time.sleep(2)

pyautogui.press("enter")
time.sleep(4)

print("Demo completed successfully!")
'''
'''
import pyautogui
import time

# Emergency stop
pyautogui.FAILSAFE = True

time.sleep(2)


# Capture the content of the webpage and save it to a Word document using PyAutoGUI
# STEP 1: OPEN CHROME
# --------------------------------------------------

print("Step 1: Open Chrome...")

pyautogui.hotkey("command", "space", interval=0.1)
time.sleep(1)

pyautogui.write("chrome", interval=0.15)
time.sleep(1)

pyautogui.press("enter")
time.sleep(4)


# --------------------------------------------------
# STEP 2: OPEN NEW TAB
# --------------------------------------------------

print("Step 2: Open a new tab...")

pyautogui.hotkey("command", "t", interval=0.1)
time.sleep(2)


# --------------------------------------------------
# STEP 3: OPEN WEBSITE
# --------------------------------------------------

print("Step 3: Go to the website...")

pyautogui.hotkey("command", "l", interval=0.1)
time.sleep(1)

pyautogui.write(
    "https://www.accuweather.com/en/in/chennai/206671/current-weather/206671",
    interval=0.03
)

time.sleep(1)

pyautogui.press("enter")
time.sleep(7)


# --------------------------------------------------
# STEP 4: SELECT WEBPAGE
# --------------------------------------------------

print("Step 4: Select the webpage...")

pyautogui.click(900, 500)
time.sleep(1)

pyautogui.hotkey("command", "a", interval=0.1)
time.sleep(1)


# --------------------------------------------------
# STEP 5: COPY WEBPAGE
# --------------------------------------------------

print("Step 5: Copy the webpage...")

pyautogui.hotkey("command", "c", interval=0.1)
time.sleep(2)


# --------------------------------------------------
# STEP 6: OPEN MICROSOFT WORD
# --------------------------------------------------

print("Step 6: Open Microsoft Word...")

pyautogui.hotkey("command", "space", interval=0.1)
time.sleep(1)

pyautogui.write(
    "Microsoft Word",
    interval=0.15
)

time.sleep(1)

pyautogui.press("enter")
time.sleep(6)


# --------------------------------------------------
# STEP 7: CREATE NEW DOCUMENT
# --------------------------------------------------

print("Step 7: Create a new Word document...")

pyautogui.hotkey(
    "command",
    "n",
    interval=0.1
)

time.sleep(3)


# --------------------------------------------------
# STEP 8: PASTE CONTENT
# --------------------------------------------------

print("Step 8: Paste the copied content...")

pyautogui.hotkey(
    "command",
    "v",
    interval=0.1
)

time.sleep(4)


# --------------------------------------------------
# STEP 9: SAVE WORD DOCUMENT
# --------------------------------------------------

print("Step 9: Save the Word document...")

pyautogui.hotkey(
    "command",
    "s",
    interval=0.1
)

time.sleep(4)

pyautogui.write(
    "pygui_demo.docx",
    interval=0.15
)

time.sleep(1)

# Go to Desktop
pyautogui.hotkey(
    "command",
    "shift",
    "d",
    interval=0.1
)

time.sleep(2)

pyautogui.press("enter")
time.sleep(5)


print("Demo completed successfully!")
'''

# Capture the content of the NSE side and save it to a Excel document using PyAutoGUI
import pyautogui
import time
import subprocess
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook


# --------------------------------------------------
# SETTINGS
# --------------------------------------------------

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3

today = datetime.now().strftime("%Y-%m-%d")
now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

excel_file = os.path.expanduser(
    f"~/Desktop/daily_report_{today}.xlsx"
)

screenshot_file = os.path.expanduser(
    f"~/Desktop/daily_report_{today}.png"
)


# --------------------------------------------------
# STEP 1: OPEN CHROME
# --------------------------------------------------

print("Step 1: Opening Chrome...")

pyautogui.hotkey("command", "space", interval=0.1)
time.sleep(1)

pyautogui.write(
    "Google Chrome",
    interval=0.15
)

time.sleep(1)
pyautogui.press("enter")
time.sleep(5)


# --------------------------------------------------
# STEP 2: OPEN NSE NIFTY 50
# --------------------------------------------------

print("Step 2: Opening NSE NIFTY 50...")

pyautogui.hotkey("command", "t", interval=0.1)
time.sleep(2)

pyautogui.hotkey("command", "l", interval=0.1)
time.sleep(1)

pyautogui.write(
    "https://www.nseindia.com/index-tracker/NIFTY%2050",
    interval=0.03
)

pyautogui.press("enter")
time.sleep(6)


# --------------------------------------------------
# STEP 3: GET NIFTY 50 VALUE
# --------------------------------------------------

print("Step 3: Reading NIFTY 50...")

pyautogui.click(900, 500)
time.sleep(1)

pyautogui.hotkey("command", "a")
time.sleep(1)

pyautogui.hotkey("command", "c")
time.sleep(2)

nifty_text = subprocess.run(
    ["pbpaste"],
    capture_output=True,
    text=True
).stdout

nifty_value = "Not detected"

match = re.search(
    r"NIFTY\s*50.*?([\d,]+\.\d+)",
    nifty_text,
    re.IGNORECASE | re.DOTALL
)

if match:
    nifty_value = match.group(1)

print(f"NIFTY 50: {nifty_value}")


# --------------------------------------------------
# STEP 4: OPEN RELIANCE DIRECTLY
# --------------------------------------------------

print("Step 4: Opening Reliance Industries...")

pyautogui.hotkey("command", "l")
time.sleep(1)

pyautogui.write(
    "https://www.nseindia.com/get-quotes/equity?symbol=RELIANCE",
    interval=0.03
)

pyautogui.press("enter")
time.sleep(6)


# --------------------------------------------------
# STEP 5: GET RELIANCE VALUE
# --------------------------------------------------

print("Step 5: Reading Reliance...")

pyautogui.click(900, 500)
time.sleep(1)

pyautogui.hotkey("command", "a")
time.sleep(1)

pyautogui.hotkey("command", "c")
time.sleep(2)

reliance_text = subprocess.run(
    ["pbpaste"],
    capture_output=True,
    text=True
).stdout

reliance_value = "Not detected"

match = re.search(
    r"Reliance Industries Limited.*?([\d,]+\.\d+)",
    reliance_text,
    re.IGNORECASE | re.DOTALL
)

if match:
    reliance_value = match.group(1)

print(f"Reliance: {reliance_value}")


# --------------------------------------------------
# STEP 6: CREATE EXCEL REPORT DIRECTLY
# --------------------------------------------------

print("Step 6: Creating Excel report...")

workbook = Workbook()
sheet = workbook.active
sheet.title = "Daily Report"

# Header
sheet.append([
    "Date & Time",
    "Index / Stock",
    "Price",
    "Comment"
])

# NIFTY 50
sheet.append([
    now,
    "NIFTY 50",
    nifty_value,
    "NIFTY 50 market data collected from NSE India."
])

# Reliance
sheet.append([
    now,
    "Reliance Industries Limited",
    reliance_value,
    "Reliance Industries market data collected from NSE India."
])

# Column widths
sheet.column_dimensions["A"].width = 22
sheet.column_dimensions["B"].width = 32
sheet.column_dimensions["C"].width = 15
sheet.column_dimensions["D"].width = 50

# Save workbook
workbook.save(excel_file)

print(f"Excel created: {excel_file}")


# --------------------------------------------------
# STEP 7: VERIFY EXCEL CONTENT
# --------------------------------------------------

print("Step 7: Verifying Excel data...")

check_book = load_workbook(
    excel_file,
    data_only=True
)

check_sheet = check_book.active

print(
    f"NIFTY 50 in Excel: {check_sheet['C2'].value}"
)

print(
    f"Reliance in Excel: {check_sheet['C3'].value}"
)

check_book.close()

if not os.path.exists(excel_file):
    raise RuntimeError(
        "Excel file was not created."
    )

print("✓ Excel file verified.")


# --------------------------------------------------
# STEP 8: OPEN THE CREATED EXCEL FILE
# --------------------------------------------------

print("Step 8: Opening the Excel report...")

subprocess.run([
    "open",
    "-a",
    "Microsoft Excel",
    excel_file
])

time.sleep(8)


# --------------------------------------------------
# STEP 9: BRING EXCEL TO FRONT
# --------------------------------------------------

print("Step 9: Bringing Excel to front...")

subprocess.run([
    "osascript",
    "-e",
    'tell application "Microsoft Excel" to activate'
])

time.sleep(5)


# --------------------------------------------------
# STEP 10: TAKE SCREENSHOT
# --------------------------------------------------

print("Step 10: Taking screenshot...")

subprocess.run([
    "screencapture",
    "-x",
    screenshot_file
])

time.sleep(2)


# --------------------------------------------------
# STEP 11: FINAL VALIDATION
# --------------------------------------------------

print("\n======================================")
print("       DAILY REPORT COMPLETED")
print("======================================")

print(f"NIFTY 50    : {nifty_value}")
print(f"Reliance    : {reliance_value}")
print(f"Excel       : {excel_file}")
print(f"Screenshot  : {screenshot_file}")

if os.path.exists(excel_file):
    print("✓ Excel file exists.")
else:
    print("✗ Excel file missing.")

if os.path.exists(screenshot_file):
    print("✓ Screenshot exists.")
else:
    print("✗ Screenshot missing.")

print("======================================")